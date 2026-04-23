"""XLSX equivalent of the readable CSV: same rows/columns, but URL cells
are real clickable hyperlinks (Font blue + underline) so users can
ctrl/cmd-click them in Excel / Numbers / LibreOffice.
"""

import logging
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

COLUMNS = [
    "city_fr",
    "hackathon_title",
    "hackathon_url",
    "hackathon_platform",
    "hackathon_location",
    "hackathon_start_date",
    "hackathon_end_date",
    "hackathon_prize_pool",
    "hackathon_score",
    "hackathon_ai_keywords",
    "hackathon_tools",
    "is_winner",
    "winner_labels",
    "project_title",
    "project_url",
    "tagline",
    "project_github",
    "built_with",
    "external_links",
    "participants",
    "participant_githubs",
    "participant_linkedins",
    "participant_twitters",
    "participant_websites",
    "participant_devpost_profiles",
]

SINGLE_URL_COLS = {"hackathon_url", "project_url", "project_github"}
MULTI_URL_COLS = {
    "external_links",
    "participant_githubs",
    "participant_linkedins",
    "participant_twitters",
    "participant_websites",
    "participant_devpost_profiles",
}

COL_WIDTHS = {
    "city_fr": 12,
    "hackathon_title": 40,
    "hackathon_url": 45,
    "hackathon_platform": 14,
    "hackathon_location": 28,
    "hackathon_start_date": 14,
    "hackathon_end_date": 14,
    "hackathon_prize_pool": 18,
    "hackathon_score": 10,
    "hackathon_ai_keywords": 20,
    "hackathon_tools": 18,
    "is_winner": 10,
    "winner_labels": 20,
    "project_title": 35,
    "project_url": 45,
    "tagline": 50,
    "project_github": 45,
    "built_with": 30,
    "external_links": 40,
    "participants": 35,
    "participant_githubs": 45,
    "participant_linkedins": 45,
    "participant_twitters": 35,
    "participant_websites": 35,
    "participant_devpost_profiles": 45,
}


def _join(values: Iterable[str], sep: str = ", ") -> str:
    return sep.join(v for v in values if v)


def _row_from_project(h: Dict, p: Dict) -> Dict:
    parts = p.get("participants") or []
    return {
        "city_fr": h.get("city_fr") or "",
        "hackathon_title": h.get("title") or "",
        "hackathon_url": h.get("url") or "",
        "hackathon_platform": h.get("source_platform") or "",
        "hackathon_location": h.get("location") or "",
        "hackathon_start_date": h.get("start_date") or "",
        "hackathon_end_date": h.get("end_date") or "",
        "hackathon_prize_pool": h.get("prize_pool") or "",
        "hackathon_score": h.get("relevance_score"),
        "hackathon_ai_keywords": _join(h.get("ai_keywords_found") or []),
        "hackathon_tools": _join(h.get("tools_found") or []),
        "is_winner": bool(p.get("is_winner")),
        "winner_labels": _join(p.get("winner_labels") or []),
        "project_title": p.get("project_title") or "",
        "project_url": p.get("project_url") or "",
        "tagline": p.get("tagline") or "",
        "project_github": p.get("github_url") or "",
        "built_with": _join(p.get("built_with") or []),
        "external_links": _join(p.get("external_links") or []),
        "participants": _join(x.get("name", "") for x in parts),
        "participant_githubs": _join(x.get("github", "") for x in parts),
        "participant_linkedins": _join(x.get("linkedin", "") for x in parts),
        "participant_twitters": _join(x.get("twitter", "") for x in parts),
        "participant_websites": _join(x.get("website", "") for x in parts),
        "participant_devpost_profiles": _join(x.get("profile_url", "") for x in parts),
    }


def _empty_row(h: Dict) -> Dict:
    return {
        "city_fr": h.get("city_fr") or "",
        "hackathon_title": h.get("title") or "",
        "hackathon_url": h.get("url") or "",
        "hackathon_platform": h.get("source_platform") or "",
        "hackathon_location": h.get("location") or "",
        "hackathon_start_date": h.get("start_date") or "",
        "hackathon_end_date": h.get("end_date") or "",
        "hackathon_prize_pool": h.get("prize_pool") or "",
        "hackathon_score": h.get("relevance_score"),
        "hackathon_ai_keywords": _join(h.get("ai_keywords_found") or []),
        "hackathon_tools": _join(h.get("tools_found") or []),
        "is_winner": "",
        "winner_labels": "",
        "project_title": "",
        "project_url": "",
        "tagline": "",
        "project_github": "",
        "built_with": "",
        "external_links": "",
        "participants": "",
        "participant_githubs": "",
        "participant_linkedins": "",
        "participant_twitters": "",
        "participant_websites": "",
        "participant_devpost_profiles": "",
    }


def write_readable_xlsx(
    hackathons: List[Dict],
    projects: Iterable[Dict],
    out_path: Path,
    winners_only: bool = False,
) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)

    by_hackathon: Dict[str, List[Dict]] = defaultdict(list)
    for p in projects:
        by_hackathon[p.get("hackathon_id") or ""].append(p)

    hackathons_sorted = sorted(
        hackathons,
        key=lambda h: (h.get("relevance_score") or 0, len(by_hackathon.get(h.get("id") or "", []))),
        reverse=True,
    )

    rows: List[Dict] = []
    for h in hackathons_sorted:
        ps = by_hackathon.get(h.get("id") or "", [])
        if winners_only:
            ps = [p for p in ps if p.get("is_winner")]
        else:
            winners = [p for p in ps if p.get("is_winner")]
            if winners:
                ps = winners + [p for p in ps if not p.get("is_winner")]
        if not ps:
            rows.append(_empty_row(h))
            continue
        for p in ps:
            rows.append(_row_from_project(h, p))

    wb = Workbook()
    ws = wb.active
    ws.title = "France"
    ws.freeze_panes = "A2"

    header_fill = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")
    link_font = Font(color="FF0563C1", underline="single")
    winner_fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 20)

    for row_idx, row in enumerate(rows, start=2):
        is_win = bool(row.get("is_winner"))
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if col_name in SINGLE_URL_COLS and value:
                cell.hyperlink = str(value)
                cell.font = link_font
            if is_win:
                cell.fill = winner_fill

    wb.save(out_path)
    logger.info("wrote readable XLSX (%d rows) to %s", len(rows), out_path)
    return out_path
